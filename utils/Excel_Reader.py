from openpyxl import load_workbook

filename = "C:\\Users\\vidhy\\PycharmProject\\GuviTask15_final\\TestData\\usrlogin_testdata.xlsx"
sheetname = "Test_data"
@staticmethod
def read_data():
    workbook = load_workbook(filename)
    sheet = workbook.active
    print("Max row:", sheet.max_row)
    data = []
    for row in range(2,sheet.max_row+1):
        username = sheet.cell(row=row, column=2).value
        password = sheet.cell(row=row, column=3).value
        data.append((row,username,password))
    return data

@staticmethod
def write_excel_data(row,result):
    workbook = load_workbook(filename)
    sheet = workbook.active
    sheet.cell(row=row, column=8).value = result
    workbook.save(filename)


